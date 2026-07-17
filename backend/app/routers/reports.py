from datetime import datetime, date, time
import calendar
from io import BytesIO
from typing import Optional

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query
)

from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill
)
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session

from ..database import get_db

from ..models import (
    Attendance,
    User,
    Leave,
    Workplace,
    Team
)

from ..security import admin_required


router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)


# =========================================================
# YARDIMCI FONKSİYONLAR
# =========================================================


def get_month_range(
    year: int,
    month: int
):

    if month < 1 or month > 12:

        raise HTTPException(
            status_code=400,
            detail="Ay değeri 1 ile 12 arasında olmalıdır."
        )

    try:

        start_date = datetime(
            year,
            month,
            1
        )

    except ValueError:

        raise HTTPException(
            status_code=400,
            detail="Geçersiz yıl veya ay bilgisi."
        )

    last_day = calendar.monthrange(
        year,
        month
    )[1]

    end_date = datetime(
        year,
        month,
        last_day,
        23,
        59,
        59
    )

    return start_date, end_date


def get_date_range(
    start_date: Optional[date],
    end_date: Optional[date]
):

    if start_date and end_date:

        if start_date > end_date:

            raise HTTPException(
                status_code=400,
                detail="Başlangıç tarihi bitiş tarihinden sonra olamaz."
            )

    start_datetime = None
    end_datetime = None

    if start_date:

        start_datetime = datetime.combine(
            start_date,
            time.min
        )

    if end_date:

        end_datetime = datetime.combine(
            end_date,
            time.max
        )

    return start_datetime, end_datetime


def calculate_duration_minutes(
    attendance: Attendance
):

    if not attendance.check_in_time:

        return 0

    end_time = (
        attendance.check_out_time
        if attendance.check_out_time
        else datetime.now()
    )

    difference = (
        end_time -
        attendance.check_in_time
    )

    return max(
        0,
        int(
            difference.total_seconds() // 60
        )
    )


def format_duration(
    total_minutes: int
):

    total_minutes = max(
        0,
        int(total_minutes or 0)
    )

    hours = total_minutes // 60
    minutes = total_minutes % 60

    return (
        f"{hours} saat "
        f"{minutes} dakika"
    )


def get_attendance_status(
    attendance: Attendance
):

    if not attendance.check_out_time:

        return "working"

    if (
        attendance.missing_minutes or 0
    ) > 0:

        return "missing"

    if attendance.late:

        return "late"

    if (
        attendance.overtime_minutes or 0
    ) > 0:

        return "overtime"

    return "completed"


def get_status_text(
    status: str
):

    status_names = {

        "working":
        "Çalışıyor",

        "missing":
        "Eksik çalışma",

        "late":
        "Geç giriş",

        "overtime":
        "Fazla mesai",

        "completed":
        "Tamamlandı"

    }

    return status_names.get(
        status,
        "Bilinmiyor"
    )


def style_excel_header(
    worksheet,
    row_number: int
):

    fill = PatternFill(
        fill_type="solid",
        fgColor="1D4ED8"
    )

    font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in worksheet[row_number]:

        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


def auto_size_columns(
    worksheet
):

    for column_cells in worksheet.columns:

        maximum_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            try:

                cell_length = len(
                    str(
                        cell.value
                        if cell.value is not None
                        else ""
                    )
                )

                maximum_length = max(
                    maximum_length,
                    cell_length
                )

            except Exception:

                pass

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            maximum_length + 3,
            40
        )


def build_admin_attendance_query(
    db: Session,

    user_id: Optional[int] = None,

    workplace_id: Optional[int] = None,

    team_id: Optional[int] = None,

    start_date: Optional[date] = None,

    end_date: Optional[date] = None
):

    query = db.query(
        Attendance,
        User,
        Workplace,
        Team
    ).join(

        User,
        Attendance.user_id == User.id

    ).outerjoin(

        Workplace,
        User.workplace_id == Workplace.id

    ).outerjoin(

        Team,
        User.team_id == Team.id

    )

    if user_id is not None:

        query = query.filter(
            User.id == user_id
        )

    if workplace_id is not None:

        query = query.filter(
            User.workplace_id == workplace_id
        )

    if team_id is not None:

        query = query.filter(
            User.team_id == team_id
        )

    start_datetime, end_datetime = (
        get_date_range(
            start_date,
            end_date
        )
    )

    if start_datetime:

        query = query.filter(
            Attendance.check_in_time >=
            start_datetime
        )

    if end_datetime:

        query = query.filter(
            Attendance.check_in_time <=
            end_datetime
        )

    return query


# =========================================================
# ADMIN FİLTRELİ MESAİ RAPORU
# =========================================================


@router.get("/admin/attendance")
def admin_attendance_report(

    user_id: Optional[int] = Query(
        default=None
    ),

    workplace_id: Optional[int] = Query(
        default=None
    ),

    team_id: Optional[int] = Query(
        default=None
    ),

    status: Optional[str] = Query(
        default=None
    ),

    start_date: Optional[date] = Query(
        default=None
    ),

    end_date: Optional[date] = Query(
        default=None
    ),

    skip: int = Query(
        default=0,
        ge=0
    ),

    limit: int = Query(
        default=100,
        ge=1,
        le=500
    ),

    db: Session = Depends(get_db),

    admin: User = Depends(admin_required)

):

    valid_statuses = {
        "working",
        "completed",
        "late",
        "overtime",
        "missing"
    }

    if (
        status and
        status not in valid_statuses
    ):

        raise HTTPException(
            status_code=400,
            detail="Geçersiz mesai durumu."
        )

    query = build_admin_attendance_query(

        db=db,

        user_id=user_id,

        workplace_id=workplace_id,

        team_id=team_id,

        start_date=start_date,

        end_date=end_date

    )

    rows = query.order_by(

        Attendance.check_in_time.desc()

    ).all()

    records = []

    total_work_minutes = 0
    total_late_minutes = 0
    total_overtime_minutes = 0
    total_missing_minutes = 0
    active_count = 0

    for (
        attendance,
        user,
        workplace,
        team
    ) in rows:

        attendance_status = (
            get_attendance_status(
                attendance
            )
        )

        if (
            status and
            attendance_status != status
        ):

            continue

        duration_minutes = (
            calculate_duration_minutes(
                attendance
            )
        )

        late_minutes = int(
            attendance.late_minutes or 0
        )

        overtime_minutes = int(
            attendance.overtime_minutes or 0
        )

        missing_minutes = int(
            attendance.missing_minutes or 0
        )

        total_work_minutes += (
            duration_minutes
        )

        total_late_minutes += (
            late_minutes
        )

        total_overtime_minutes += (
            overtime_minutes
        )

        total_missing_minutes += (
            missing_minutes
        )

        if attendance_status == "working":

            active_count += 1

        records.append({

            "id":
            attendance.id,

            "user_id":
            user.id,

            "personel":
            f"{user.name} {user.surname}",

            "email":
            user.email,

            "role":
            user.role,

            "workplace_id":
            user.workplace_id,

            "workplace":
            workplace.name
            if workplace
            else None,

            "team_id":
            user.team_id,

            "team":
            team.name
            if team
            else None,

            "check_in":
            attendance.check_in_time,

            "check_out":
            attendance.check_out_time,

            "duration_minutes":
            duration_minutes,

            "duration":
            format_duration(
                duration_minutes
            ),

            "late":
            bool(attendance.late),

            "late_minutes":
            late_minutes,

            "overtime_minutes":
            overtime_minutes,

            "missing_minutes":
            missing_minutes,

            "status":
            attendance_status,

            "status_text":
            get_status_text(
                attendance_status
            )

        })

    total_records = len(records)

    paginated_records = records[
        skip:
        skip + limit
    ]

    return {

        "filters": {

            "user_id":
            user_id,

            "workplace_id":
            workplace_id,

            "team_id":
            team_id,

            "status":
            status,

            "start_date":
            start_date,

            "end_date":
            end_date

        },

        "summary": {

            "total_records":
            total_records,

            "active_count":
            active_count,

            "total_work_minutes":
            total_work_minutes,

            "total_work":
            format_duration(
                total_work_minutes
            ),

            "total_late_minutes":
            total_late_minutes,

            "total_overtime_minutes":
            total_overtime_minutes,

            "total_missing_minutes":
            total_missing_minutes

        },

        "pagination": {

            "skip":
            skip,

            "limit":
            limit,

            "returned":
            len(paginated_records),

            "total":
            total_records

        },

        "records":
        paginated_records

    }


# =========================================================
# ADMIN FİLTRELİ EXCEL RAPORU
# =========================================================


@router.get("/admin/attendance/excel")
def admin_attendance_excel(

    user_id: Optional[int] = Query(
        default=None
    ),

    workplace_id: Optional[int] = Query(
        default=None
    ),

    team_id: Optional[int] = Query(
        default=None
    ),

    status: Optional[str] = Query(
        default=None
    ),

    start_date: Optional[date] = Query(
        default=None
    ),

    end_date: Optional[date] = Query(
        default=None
    ),

    db: Session = Depends(get_db),

    admin: User = Depends(admin_required)

):

    valid_statuses = {
        "working",
        "completed",
        "late",
        "overtime",
        "missing"
    }

    if (
        status and
        status not in valid_statuses
    ):

        raise HTTPException(
            status_code=400,
            detail="Geçersiz mesai durumu."
        )

    query = build_admin_attendance_query(

        db=db,

        user_id=user_id,

        workplace_id=workplace_id,

        team_id=team_id,

        start_date=start_date,

        end_date=end_date

    )

    rows = query.order_by(

        Attendance.check_in_time.desc()

    ).all()

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Mesai Raporu"

    worksheet.merge_cells(
        "A1:L1"
    )

    worksheet["A1"] = (
        "PDKS KURUMSAL MESAİ RAPORU"
    )

    worksheet["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    worksheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="111827"
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    worksheet.append([])

    worksheet.append([
        "Personel",
        "E-posta",
        "İş Yeri",
        "Ekip",
        "Giriş",
        "Çıkış",
        "Çalışma Süresi",
        "Geç Kalma",
        "Fazla Mesai",
        "Eksik Çalışma",
        "Durum",
        "Kayıt No"
    ])

    style_excel_header(
        worksheet,
        3
    )

    total_work_minutes = 0
    total_late_minutes = 0
    total_overtime_minutes = 0
    total_missing_minutes = 0
    exported_count = 0

    for (
        attendance,
        user,
        workplace,
        team
    ) in rows:

        attendance_status = (
            get_attendance_status(
                attendance
            )
        )

        if (
            status and
            attendance_status != status
        ):

            continue

        duration_minutes = (
            calculate_duration_minutes(
                attendance
            )
        )

        late_minutes = int(
            attendance.late_minutes or 0
        )

        overtime_minutes = int(
            attendance.overtime_minutes or 0
        )

        missing_minutes = int(
            attendance.missing_minutes or 0
        )

        total_work_minutes += (
            duration_minutes
        )

        total_late_minutes += (
            late_minutes
        )

        total_overtime_minutes += (
            overtime_minutes
        )

        total_missing_minutes += (
            missing_minutes
        )

        exported_count += 1

        worksheet.append([

            f"{user.name} {user.surname}",

            user.email,

            workplace.name
            if workplace
            else "Atanmamış",

            team.name
            if team
            else "Atanmamış",

            attendance.check_in_time,

            attendance.check_out_time,

            format_duration(
                duration_minutes
            ),

            late_minutes,

            overtime_minutes,

            missing_minutes,

            get_status_text(
                attendance_status
            ),

            attendance.id

        ])

    worksheet.append([])

    summary_row = worksheet.max_row + 1

    worksheet.append([
        "ÖZET",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        ""
    ])

    worksheet.merge_cells(
        start_row=summary_row,
        start_column=1,
        end_row=summary_row,
        end_column=12
    )

    worksheet.cell(
        row=summary_row,
        column=1
    ).font = Font(
        bold=True,
        color="FFFFFF"
    )

    worksheet.cell(
        row=summary_row,
        column=1
    ).fill = PatternFill(
        fill_type="solid",
        fgColor="374151"
    )

    worksheet.append([
        "Toplam Kayıt",
        exported_count
    ])

    worksheet.append([
        "Toplam Çalışma",
        format_duration(
            total_work_minutes
        )
    ])

    worksheet.append([
        "Toplam Geç Kalma",
        f"{total_late_minutes} dakika"
    ])

    worksheet.append([
        "Toplam Fazla Mesai",
        f"{total_overtime_minutes} dakika"
    ])

    worksheet.append([
        "Toplam Eksik Çalışma",
        f"{total_missing_minutes} dakika"
    ])

    auto_size_columns(
        worksheet
    )

    file = BytesIO()

    workbook.save(file)

    file.seek(0)

    file_name = (
        "pdks_admin_mesai_raporu_"
        f"{datetime.now().strftime('%Y%m%d_%H%M')}"
        ".xlsx"
    )

    return StreamingResponse(

        file,

        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),

        headers={

            "Content-Disposition":
            f'attachment; filename="{file_name}"'

        }

    )


# =========================================================
# TÜM PERSONEL AYLIK RAPOR
# Statik endpoint dinamik endpointten önce tanımlandı.
# =========================================================


@router.get("/monthly/all")
def all_monthly_report(

    year: int,

    month: int,

    db: Session = Depends(get_db),

    admin: User = Depends(admin_required)

):

    start_date, end_date = (
        get_month_range(
            year,
            month
        )
    )

    users = db.query(User).filter(
        User.role == "employee"
    ).order_by(
        User.name.asc(),
        User.surname.asc()
    ).all()

    result = []

    for user in users:

        records = db.query(
            Attendance
        ).filter(

            Attendance.user_id ==
            user.id,

            Attendance.check_in_time >=
            start_date,

            Attendance.check_in_time <=
            end_date

        ).all()

        total_minutes = 0
        overtime = 0
        missing = 0
        late = 0
        late_count = 0

        for record in records:

            total_minutes += (
                calculate_duration_minutes(
                    record
                )
            )

            if record.late:

                late_count += 1

            late += int(
                record.late_minutes or 0
            )

            overtime += int(
                record.overtime_minutes or 0
            )

            missing += int(
                record.missing_minutes or 0
            )

        workplace = None
        team = None

        if user.workplace_id:

            workplace = db.query(
                Workplace
            ).filter(

                Workplace.id ==
                user.workplace_id

            ).first()

        if user.team_id:

            team = db.query(
                Team
            ).filter(

                Team.id ==
                user.team_id

            ).first()

        result.append({

            "user_id":
            user.id,

            "personel":
            f"{user.name} {user.surname}",

            "workplace_id":
            user.workplace_id,

            "workplace":
            workplace.name
            if workplace
            else None,

            "team_id":
            user.team_id,

            "team":
            team.name
            if team
            else None,

            "calisma_gunu":
            len(records),

            "toplam_calisma_dakika":
            total_minutes,

            "toplam_saat":
            format_duration(
                total_minutes
            ),

            "gecikme_sayisi":
            late_count,

            "gecikme_dakika":
            late,

            "fazla_mesai":
            overtime,

            "eksik_mesai":
            missing

        })

    return result


# =========================================================
# TEK PERSONEL AYLIK RAPOR
# =========================================================


@router.get("/monthly/{user_id}")
def monthly_report(

    user_id: int,

    year: int,

    month: int,

    db: Session = Depends(get_db),

    admin: User = Depends(admin_required)

):

    user = db.query(User).filter(
        User.id == user_id
    ).first()

    if not user:

        raise HTTPException(
            status_code=404,
            detail="Personel bulunamadı"
        )

    start_date, end_date = (
        get_month_range(
            year,
            month
        )
    )

    attendances = db.query(
        Attendance
    ).filter(

        Attendance.user_id ==
        user_id,

        Attendance.check_in_time >=
        start_date,

        Attendance.check_in_time <=
        end_date

    ).order_by(

        Attendance.check_in_time.asc()

    ).all()

    total_minutes = 0
    late_count = 0
    late_minutes = 0
    total_overtime = 0
    total_missing = 0

    records = []

    for record in attendances:

        duration_minutes = (
            calculate_duration_minutes(
                record
            )
        )

        total_minutes += (
            duration_minutes
        )

        if record.late:

            late_count += 1

            late_minutes += int(
                record.late_minutes or 0
            )

        total_overtime += int(
            record.overtime_minutes or 0
        )

        total_missing += int(
            record.missing_minutes or 0
        )

        status = get_attendance_status(
            record
        )

        records.append({

            "id":
            record.id,

            "check_in":
            record.check_in_time,

            "check_out":
            record.check_out_time,

            "duration_minutes":
            duration_minutes,

            "duration":
            format_duration(
                duration_minutes
            ),

            "late":
            bool(record.late),

            "late_minutes":
            int(
                record.late_minutes or 0
            ),

            "overtime_minutes":
            int(
                record.overtime_minutes or 0
            ),

            "missing_minutes":
            int(
                record.missing_minutes or 0
            ),

            "status":
            status,

            "status_text":
            get_status_text(
                status
            )

        })

    leaves = db.query(Leave).filter(

        Leave.user_id == user_id,

        Leave.status == "approved",

        Leave.start_date <=
        end_date.date(),

        Leave.end_date >=
        start_date.date()

    ).all()

    leave_days = 0

    for leave in leaves:

        leave_start = max(
            leave.start_date,
            start_date.date()
        )

        leave_end = min(
            leave.end_date,
            end_date.date()
        )

        leave_days += (
            leave_end -
            leave_start
        ).days + 1

    workplace = None
    team = None

    if user.workplace_id:

        workplace = db.query(
            Workplace
        ).filter(

            Workplace.id ==
            user.workplace_id

        ).first()

    if user.team_id:

        team = db.query(
            Team
        ).filter(

            Team.id ==
            user.team_id

        ).first()

    return {

        "user_id":
        user.id,

        "personel":
        f"{user.name} {user.surname}",

        "email":
        user.email,

        "workplace_id":
        user.workplace_id,

        "workplace":
        workplace.name
        if workplace
        else None,

        "team_id":
        user.team_id,

        "team":
        team.name
        if team
        else None,

        "ay":
        f"{year}-{month:02}",

        "calisma_gunu":
        len(attendances),

        "toplam_calisma_dakika":
        total_minutes,

        "toplam_saat":
        format_duration(
            total_minutes
        ),

        "gecikme_sayisi":
        late_count,

        "gecikme_dakika":
        late_minutes,

        "izin_gunu":
        leave_days,

        "fazla_mesai_dakika":
        total_overtime,

        "eksik_mesai_dakika":
        total_missing,

        "records":
        records

    }


# =========================================================
# TEK PERSONEL AYLIK EXCEL RAPORU
# =========================================================


@router.get("/monthly/excel/{user_id}")
def monthly_excel_report(

    user_id: int,

    year: int,

    month: int,

    db: Session = Depends(get_db),

    admin: User = Depends(admin_required)

):

    user = db.query(User).filter(
        User.id == user_id
    ).first()

    if not user:

        raise HTTPException(
            status_code=404,
            detail="Personel bulunamadı"
        )

    start_date, end_date = (
        get_month_range(
            year,
            month
        )
    )

    records = db.query(
        Attendance
    ).filter(

        Attendance.user_id ==
        user_id,

        Attendance.check_in_time >=
        start_date,

        Attendance.check_in_time <=
        end_date

    ).order_by(

        Attendance.check_in_time.asc()

    ).all()

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Aylık Rapor"

    worksheet.merge_cells(
        "A1:G1"
    )

    worksheet["A1"] = (
        "PDKS AYLIK PERSONEL RAPORU"
    )

    worksheet["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    worksheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="111827"
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    worksheet.append([])

    worksheet.append([
        "Personel",
        f"{user.name} {user.surname}"
    ])

    worksheet.append([
        "Dönem",
        f"{year}-{month:02}"
    ])

    worksheet.append([])

    worksheet.append([
        "Giriş",
        "Çıkış",
        "Süre",
        "Geç Kalma",
        "Fazla Mesai",
        "Eksik Çalışma",
        "Durum"
    ])

    style_excel_header(
        worksheet,
        6
    )

    total_minutes = 0
    late_count = 0
    late_minutes = 0
    overtime = 0
    missing = 0

    for record in records:

        duration_minutes = (
            calculate_duration_minutes(
                record
            )
        )

        total_minutes += (
            duration_minutes
        )

        if record.late:

            late_count += 1

        late_minutes += int(
            record.late_minutes or 0
        )

        overtime += int(
            record.overtime_minutes or 0
        )

        missing += int(
            record.missing_minutes or 0
        )

        status = get_attendance_status(
            record
        )

        worksheet.append([

            record.check_in_time,

            record.check_out_time,

            format_duration(
                duration_minutes
            ),

            int(
                record.late_minutes or 0
            ),

            int(
                record.overtime_minutes or 0
            ),

            int(
                record.missing_minutes or 0
            ),

            get_status_text(
                status
            )

        ])

    worksheet.append([])

    worksheet.append([
        "ÖZET"
    ])

    worksheet.append([
        "Toplam Çalışma",
        format_duration(
            total_minutes
        )
    ])

    worksheet.append([
        "Çalışma Günü",
        len(records)
    ])

    worksheet.append([
        "Geç Kalma Sayısı",
        late_count
    ])

    worksheet.append([
        "Geç Kalma Dakika",
        late_minutes
    ])

    worksheet.append([
        "Fazla Mesai Dakika",
        overtime
    ])

    worksheet.append([
        "Eksik Mesai Dakika",
        missing
    ])

    auto_size_columns(
        worksheet
    )

    file = BytesIO()

    workbook.save(file)

    file.seek(0)

    safe_name = (
        f"{user.name}_{user.surname}"
        .replace(" ", "_")
    )

    return StreamingResponse(

        file,

        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),

        headers={

            "Content-Disposition":
            (
                "attachment; "
                f'filename="{safe_name}_'
                f'{year}_{month:02}_rapor.xlsx"'
            )

        }

    )


# =========================================================
# İŞ YERİ BAZLI AYLIK RAPOR
# =========================================================


@router.get("/workplace/{workplace_id}")
def workplace_monthly_report(

    workplace_id: int,

    year: int,

    month: int,

    db: Session = Depends(get_db),

    admin: User = Depends(admin_required)

):

    workplace = db.query(
        Workplace
    ).filter(

        Workplace.id ==
        workplace_id

    ).first()

    if not workplace:

        raise HTTPException(
            status_code=404,
            detail="İş yeri bulunamadı"
        )

    start_date, end_date = (
        get_month_range(
            year,
            month
        )
    )

    users = db.query(User).filter(

        User.workplace_id ==
        workplace_id,

        User.role ==
        "employee"

    ).order_by(

        User.name.asc(),
        User.surname.asc()

    ).all()

    result = []

    for user in users:

        records = db.query(
            Attendance
        ).filter(

            Attendance.user_id ==
            user.id,

            Attendance.check_in_time >=
            start_date,

            Attendance.check_in_time <=
            end_date

        ).all()

        total_minutes = 0
        overtime = 0
        missing = 0
        late = 0

        for record in records:

            total_minutes += (
                calculate_duration_minutes(
                    record
                )
            )

            overtime += int(
                record.overtime_minutes or 0
            )

            missing += int(
                record.missing_minutes or 0
            )

            late += int(
                record.late_minutes or 0
            )

        team = None

        if user.team_id:

            team = db.query(
                Team
            ).filter(

                Team.id ==
                user.team_id

            ).first()

        result.append({

            "user_id":
            user.id,

            "personel":
            f"{user.name} {user.surname}",

            "workplace_id":
            workplace.id,

            "workplace":
            workplace.name,

            "team_id":
            user.team_id,

            "team":
            team.name
            if team
            else None,

            "calisma_gunu":
            len(records),

            "toplam_calisma_dakika":
            total_minutes,

            "toplam_saat":
            format_duration(
                total_minutes
            ),

            "gecikme_dakika":
            late,

            "fazla_mesai":
            overtime,

            "eksik_mesai":
            missing

        })

    return {

        "workplace_id":
        workplace.id,

        "workplace":
        workplace.name,

        "year":
        year,

        "month":
        month,

        "employee_count":
        len(users),

        "records":
        result

    }